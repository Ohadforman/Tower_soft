#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path("/Users/ohadformanair/PycharmProjects/Tower_work")
MAINT_DIR = ROOT / "maintenance"
PART_ORDERS = ROOT / "data" / "part_orders.csv"
STATE_PATH = MAINT_DIR / "_app_state.json"
DATASET_DIR = ROOT / "data_set_csv"


def safe_str(v) -> str:
    return "" if v is None else str(v)


def mode_norm(x: str) -> str:
    s = safe_str(x).strip().lower()
    if s in ("draw", "draws", "draws_count", "draw_count"):
        return "draws"
    return s


def norm_source(s: str) -> str:
    return safe_str(s).strip().lower()


def legacy_either_hours_interval(row: dict) -> float | None:
    txt = " ".join(
        [
            safe_str(row.get("Interval Unit", "")),
            safe_str(row.get("Trigger Context", "")),
            safe_str(row.get("Task Name", "")),
        ]
    ).lower()
    m = re.search(r"(\d+(?:\.\d+)?)\s*operating\s*hours", txt)
    if not m:
        m = re.search(r"(\d+(?:\.\d+)?)\s*hours", txt)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def legacy_either_calendar(row: dict) -> tuple[float | None, str | None]:
    unit_txt = " ".join(
        [
            safe_str(row.get("Interval Unit", "")),
            safe_str(row.get("Calendar Rule", "")),
            safe_str(row.get("Trigger Context", "")),
            safe_str(row.get("Task Name", "")),
        ]
    ).lower()
    try:
        value = float(row.get("Interval Value", ""))
    except Exception:
        value = None
    if value is None:
        return None, None
    for needle, unit in (("month", "months"), ("week", "weeks"), ("day", "days"), ("year", "years")):
        if needle in unit_txt:
            return value, unit
    return None, None


def split_trigger_modes(row: dict) -> list[str]:
    raw = safe_str(row.get("Trigger_Modes", "")).strip()
    out: list[str] = []
    if raw:
        for p in raw.replace(";", ",").replace("|", ",").split(","):
            mv = mode_norm(p)
            if mv in ("hours", "draws", "calendar", "event") and mv not in out:
                out.append(mv)
    if out:
        return out
    tracking_mode = safe_str(row.get("Tracking Mode", "")).strip().lower()
    if tracking_mode in ("either", "any", "both"):
        cal_value, _ = legacy_either_calendar(row)
        hours_value = legacy_either_hours_interval(row)
        if cal_value is not None:
            out.append("calendar")
        if hours_value is not None:
            out.append("hours")
        return out
    mv = mode_norm(tracking_mode)
    return [mv] if mv else []


def split_hours_sources(row: dict) -> list[str]:
    raw = safe_str(row.get("Trigger_Hours_Source", "")).strip()
    if not raw:
        raw = safe_str(row.get("Hours Source", "")).strip()
    if not raw:
        return ["furnace"]
    out: list[str] = []
    for p in raw.replace(";", ",").replace("|", ",").split(","):
        sv = norm_source(p)
        if sv in ("uv", "uv both", "uv-both", "uv1+uv2", "uv2+uv1"):
            for uv in ("uv1", "uv2"):
                if uv not in out:
                    out.append(uv)
        elif sv in ("uv1", "uv 1", "uv_system_1", "uv system 1", "uv-system-1", "system1", "system 1"):
            if "uv1" not in out:
                out.append("uv1")
        elif sv in ("uv2", "uv 2", "uv_system_2", "uv system 2", "uv-system-2", "system2", "system 2"):
            if "uv2" not in out:
                out.append("uv2")
        else:
            if "furnace" not in out:
                out.append("furnace")
    return out or ["furnace"]


def interval_days(value: float | None, unit: str | None) -> int | None:
    if value is None:
        return None
    u = safe_str(unit).strip().lower()
    mult = 1
    if "week" in u:
        mult = 7
    elif "month" in u:
        mult = 30
    elif "year" in u:
        mult = 365
    elif "day" in u:
        mult = 1
    return max(1, int(round(float(value) * mult)))


def stable_bucket(task_id: str, task_name: str) -> str:
    key = task_id or task_name or "task"
    score = sum(ord(ch) for ch in key) % 10
    if score == 0:
        return "overdue"
    if score in (1, 2, 3):
        return "due_soon"
    return "ok"


def days_since_for_bucket(interval: int, threshold: int, bucket: str) -> int:
    if bucket == "overdue":
        return interval + max(1, min(10, threshold))
    if bucket == "due_soon":
        remaining = max(1, min(max(1, threshold // 2), max(1, int(interval * 0.15))))
        return max(0, interval - remaining)
    remaining = max(threshold + 5, int(interval * 0.35))
    remaining = min(max(1, interval - 2), remaining)
    return max(0, interval - remaining)


def hours_since_for_bucket(interval: float, warn_hours: float, bucket: str) -> float:
    if bucket == "overdue":
        return float(interval + max(5.0, min(25.0, warn_hours)))
    if bucket == "due_soon":
        remaining = max(5.0, min(max(5.0, warn_hours * 0.5), max(5.0, interval * 0.2)))
        return max(0.0, float(interval - remaining))
    remaining = max(warn_hours + 10.0, float(interval) * 0.35)
    remaining = min(max(5.0, float(interval) - 5.0), remaining)
    return max(0.0, float(interval - remaining))


def draws_since_for_bucket(interval: int, bucket: str) -> int:
    if bucket == "overdue":
        return interval + 2
    if bucket == "due_soon":
        return max(0, interval - 2)
    remaining = min(max(2, interval - 2), max(8, int(interval * 0.35)))
    return max(0, interval - remaining)


def ensure_column(ws, header_map: dict[str, int], name: str) -> int:
    if name in header_map:
        return header_map[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    header_map[name] = col
    return col


def read_row(ws, row_idx: int, header_map: dict[str, int]) -> dict:
    return {name: ws.cell(row=row_idx, column=col).value for name, col in header_map.items()}


def write_value(ws, row_idx: int, header_map: dict[str, int], name: str, value) -> None:
    col = ensure_column(ws, header_map, name)
    ws.cell(row=row_idx, column=col).value = value


def current_hours_for_source(source: str, furnace: float, uv1: float, uv2: float) -> float:
    if source == "uv1":
        return uv1
    if source == "uv2":
        return uv2
    return furnace


def main() -> None:
    now_dt = datetime.now()
    current_date = date.today()

    state = {}
    if STATE_PATH.exists():
        state = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    furnace_hours = float(state.get("furnace_hours", 1010.0) or 1010.0)
    uv1_hours = float(state.get("uv1_hours", 250.0) or 250.0)
    uv2_hours = float(state.get("uv2_hours", 250.0) or 250.0)
    current_draw_count = sum(1 for _ in DATASET_DIR.glob("*.csv"))
    warn_days = int(state.get("warn_days", 14) or 14)
    warn_hours = float(state.get("warn_hours", 50.0) or 50.0)

    backup_dir = MAINT_DIR / f"_realistic_baseline_backup_{now_dt.strftime('%Y%m%d_%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    for path in sorted(MAINT_DIR.glob("*_Maintenance_Tracker_Template.xlsx")):
        shutil.copy2(path, backup_dir / path.name)
    if STATE_PATH.exists():
        shutil.copy2(STATE_PATH, backup_dir / STATE_PATH.name)
    for extra in [MAINT_DIR / "maintenance_task_state.csv", MAINT_DIR / "maintenance_wait_parts_log.csv", PART_ORDERS]:
        if extra.exists():
            shutil.copy2(extra, backup_dir / extra.name)

    updated_tasks = 0
    for path in sorted(MAINT_DIR.glob("*_Maintenance_Tracker_Template.xlsx")):
        wb = load_workbook(path)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_map = {safe_str(v).strip(): i for i, v in enumerate(headers, start=1) if safe_str(v).strip()}
        ensure_column(ws, header_map, "Last Done Date")
        ensure_column(ws, header_map, "Last Done Hours")
        ensure_column(ws, header_map, "Last Done Furnace Hours")
        ensure_column(ws, header_map, "Last_Done_Draw")
        ensure_column(ws, header_map, "Last_Done_Hours_UV1")
        ensure_column(ws, header_map, "Last_Done_Hours_UV2")
        ensure_column(ws, header_map, "Last_Done_Hours_Furnace")

        for row_idx in range(2, ws.max_row + 1):
            row = read_row(ws, row_idx, header_map)
            task_id = safe_str(row.get("Task ID", "")).strip()
            task_name = safe_str(row.get("Task Name", "")).strip()
            if not task_id and not task_name:
                continue
            bucket = stable_bucket(task_id, task_name)
            modes = split_trigger_modes(row)
            tracking_mode = mode_norm(row.get("Tracking Mode", ""))

            if tracking_mode == "event" or modes == ["event"]:
                write_value(ws, row_idx, header_map, "Last Done Date", current_date.isoformat())
                write_value(ws, row_idx, header_map, "Last_Done_Draw", max(0, current_draw_count - 1))
                srcs = split_hours_sources(row)
                primary_src = srcs[0] if srcs else "furnace"
                primary_cur = current_hours_for_source(primary_src, furnace_hours, uv1_hours, uv2_hours)
                write_value(ws, row_idx, header_map, "Last Done Hours", float(max(0.0, primary_cur - 1.0)))
                write_value(ws, row_idx, header_map, "Last Done Furnace Hours", float(max(0.0, furnace_hours - 1.0)))
                write_value(ws, row_idx, header_map, "Last_Done_Hours_Furnace", float(max(0.0, furnace_hours - 1.0)))
                write_value(ws, row_idx, header_map, "Last_Done_Hours_UV1", float(max(0.0, uv1_hours - 1.0)))
                write_value(ws, row_idx, header_map, "Last_Done_Hours_UV2", float(max(0.0, uv2_hours - 1.0)))
                updated_tasks += 1
                continue

            if "calendar" in modes:
                cal_value = row.get("Trigger_Calendar_Value", None)
                cal_unit = row.get("Trigger_Calendar_Unit", None)
                try:
                    cal_value = float(cal_value) if safe_str(cal_value).strip() != "" else None
                except Exception:
                    cal_value = None
                if tracking_mode in ("either", "any", "both"):
                    legacy_v, legacy_u = legacy_either_calendar(row)
                    if legacy_v is not None:
                        cal_value, cal_unit = legacy_v, legacy_u
                if cal_value is None:
                    try:
                        cal_value = float(row.get("Interval Value", ""))
                    except Exception:
                        cal_value = None
                if not safe_str(cal_unit).strip():
                    cal_unit = row.get("Interval Unit", "")
                int_days = interval_days(cal_value, cal_unit)
                if int_days:
                    threshold = int(float(row.get("Due Threshold (days)", warn_days) or warn_days))
                    days_since = days_since_for_bucket(int_days, threshold, bucket)
                    write_value(ws, row_idx, header_map, "Last Done Date", (current_date - timedelta(days=days_since)).isoformat())

            if "hours" in modes:
                raw_h = row.get("Trigger_Hours_Interval", None)
                try:
                    interval_h = float(raw_h) if safe_str(raw_h).strip() != "" else None
                except Exception:
                    interval_h = None
                if tracking_mode in ("either", "any", "both"):
                    legacy_h = legacy_either_hours_interval(row)
                    if legacy_h is not None:
                        interval_h = legacy_h
                if interval_h is None:
                    try:
                        interval_h = float(row.get("Interval Value", ""))
                    except Exception:
                        interval_h = None
                if interval_h is not None and interval_h > 0:
                    srcs = split_hours_sources(row)
                    for src in srcs:
                        cur_h = current_hours_for_source(src, furnace_hours, uv1_hours, uv2_hours)
                        last_h = max(0.0, cur_h - hours_since_for_bucket(interval_h, warn_hours, bucket))
                        if src == "uv1":
                            write_value(ws, row_idx, header_map, "Last_Done_Hours_UV1", float(last_h))
                        elif src == "uv2":
                            write_value(ws, row_idx, header_map, "Last_Done_Hours_UV2", float(last_h))
                        else:
                            write_value(ws, row_idx, header_map, "Last Done Furnace Hours", float(last_h))
                            write_value(ws, row_idx, header_map, "Last_Done_Hours_Furnace", float(last_h))
                    primary_src = srcs[0] if srcs else "furnace"
                    write_value(
                        ws,
                        row_idx,
                        header_map,
                        "Last Done Hours",
                        float(max(0.0, current_hours_for_source(primary_src, furnace_hours, uv1_hours, uv2_hours) - hours_since_for_bucket(interval_h, warn_hours, bucket))),
                    )

            if "draws" in modes:
                raw_d = row.get("Trigger_Draws_Interval", None)
                try:
                    interval_d = int(float(raw_d)) if safe_str(raw_d).strip() != "" else None
                except Exception:
                    interval_d = None
                if interval_d is None:
                    try:
                        interval_d = int(float(row.get("Interval Value", "")))
                    except Exception:
                        interval_d = None
                if interval_d is not None and interval_d > 0:
                    last_draw = max(0, int(current_draw_count - draws_since_for_bucket(interval_d, bucket)))
                    write_value(ws, row_idx, header_map, "Last_Done_Draw", last_draw)

            updated_tasks += 1

        wb.save(path)

    state["current_date"] = current_date.isoformat()
    state["furnace_hours"] = furnace_hours
    state["uv1_hours"] = uv1_hours
    state["uv2_hours"] = uv2_hours
    state["warn_days"] = warn_days
    state["warn_hours"] = warn_hours
    state["last_draw_count"] = int(current_draw_count)
    state["status_weekly_updated_at"] = now_dt.strftime("%Y-%m-%d %H:%M:%S")
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=True, indent=2), encoding="utf-8")

    # Clear old maintenance wait/test artifacts so the new baseline reflects the real timing logic.
    mts = MAINT_DIR / "maintenance_task_state.csv"
    mts.write_text(
        "task_key,task_id,component,task,state,updated_ts,updated_by,note\n",
        encoding="utf-8",
    )
    mwl = MAINT_DIR / "maintenance_wait_parts_log.csv"
    mwl.write_text(
        "wait_id,wait_ts,maintenance_component,maintenance_task,maintenance_task_id,maintenance_source_file,requested_part_name,requested_part_plan_json,requested_project_name,requested_company,wait_reason,actor,resolved_ts,resolution_note\n",
        encoding="utf-8",
    )
    if PART_ORDERS.exists():
        df_orders = pd.read_csv(PART_ORDERS).fillna("")
        if "Maintenance Task ID" in df_orders.columns:
            df_orders = df_orders[df_orders["Maintenance Task ID"].astype(str).str.strip().eq("")].copy()
            df_orders.to_csv(PART_ORDERS, index=False)

    print(
        json.dumps(
            {
                "backup_dir": str(backup_dir),
                "updated_tasks": updated_tasks,
                "current_date": state["current_date"],
                "furnace_hours": state["furnace_hours"],
                "uv1_hours": state["uv1_hours"],
                "uv2_hours": state["uv2_hours"],
                "draw_count": state["last_draw_count"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
